# -*- coding: UTF-8 -*-

from csv import reader
from os import path
from time import time
from openpyxl import Workbook
from gc import collect
from sys import exit

motd1 = '''
  _____   _    _  _____     _____  ______  ____   ______  _____
 |  __ \\ | |  | ||  __ \\   / ____||  ____|/ __ \\ |  ____||  __ \\
 | |__) || |__| || |__) | | (___  | |__  | |  | || |__   | |__) |
 |  ___/ |  __  ||  ___/   \\___ \\ |  __| | |  | ||  __|  |  _  /
 | |     | |  | || |       ____) || |____| |__| || |____ | | \\ \\
 |_|__ __|_|  |_||_|  ____|_____/ |______|\\___/ |______||_|__\\\\
   / // __ \\ |  __ \\ |  ____|| \\ | ||  __ \\ | |  | | / ____|\\ \\
  / /| |  | || |__) || |__   |  \\| || |__) || |__| || |      \\ \\
 < < | |  | ||  ___/ |  __|  | . ` ||  _  / |  __  || |       > >
  \\ \\| |__| || |     | |____ | |\\  || | \\ \\ | |  | || |____  / / 
   \\_\\\\____/ |_|     |______||_| \\_||_|  \\_\\|_|  |_| \\_____|/_/

'''

print(motd1)

###################
# 读取关键词
###################
def readKeys():
    keys = []                                     # 关键词列表

    if not path.exists('keywords.ini'):
        keys = ['么', '吗', '哪', '如何', '为何', '是否', '怎样', '还是',
        '无法', '区别', '不同', '不一样', '多少', '多久', '多大', '不',
        '乱码', '错', '失败', '几', '啥', '原因', '步骤', '流程']
        with open('keywords.ini', 'w') as f:
            f.write('\n'.join(keys))              # 写入默认关键词
        print('>>>将使用默认关键词进行词包筛选(', str(len(keys)), '个)\n', ' '.join(keys))
    else:
        with open('keywords.ini', 'r') as f:
            keys = f.read().split('\n')           # 读取关键词
        print('>>>将使用以下关键词进行词包筛选(', str(len(keys)), '个)\n', ' '.join(keys))

    return keys


###################
# 读取文件路径
###################
def readFile():
    file = {}
    while True:
        print("\n>>>提示:(*.csv)/词包路径 r/重新加载关键词 h/帮助 q/退出。")
        url = input('请输入: ')                      # 文件路径 c:\py\test.csv
        url = url.strip()                            # 去除空格
        url = url.strip('"')                         # 去除双引号 拖进来时会有双引号

        if url == 'q':
            exit()

        if url == 'h':
            print('* 请查看README.txt文件获取帮助信息。\n* 本软件更新地址:\n\tGitHub链接：https://github.com/openrhc/phpcn-seoer')
            continue

        if url == 'r':
            keys = readKeys()
            continue

        if not path.exists(url):
            print('该文件不存在！')
            continue

        basename = path.basename(url)                 # 文件名 test.csv
        suffix = path.splitext(basename)[1]           # 后缀名 .csv

        suffixs = ['.csv']                            # 支持的词包格式

        if not suffix in suffixs:
            print('请将文件转为: ', ' '.join(suffixs), ' 格式')
            continue

        file['filename'] = path.splitext(basename)[0] # 文件名 test
        file['basename'] = basename                   # 文件名 test.csv
        file['abspath']  = path.abspath(url)          # 绝对路径 f:\test.csv
        file['filedir']  = path.dirname(url)          # 目录名
        break

    return file


#####################################################
# 主函数
#####################################################

def main():

    file = readFile()

    print('\n[词包名称]  : ' + file['basename'])
    print('[词包路径]  : '   + file['abspath'])
    print('[关键词列表]: '   + ' '.join(keys), '\n')

    ipt = input('>>>确认开始筛选吗？: [Y/n] ')
    if not ipt in ('y', 'Y'):
        print('中止。')
        return

    rawSheet = []                # 原词
    filterSheet = []             # 疑问词表
    nofilterSheet = []           # 非疑问词表

    print('正在筛选...')

    startTime = time()           # 运行开始时间

    # 打开文件
    with open(file['abspath'], 'r', errors='ignore') as f:
        csv_reader = reader(f)

        lens = len(keys)              # 关键词数量
        index = 1                     # 关键词索引

        
        for line in csv_reader:
            rawSheet.append(line[0])                             # 加入未处理列表
            index = 1
            for key in keys:
                if key in line[0]:
                    # print(line[0], '... 符合关键词要求')
                    filterSheet.append(line[0])                  # 加入疑问词列表
                    break
                elif index == lens:
                    nofilterSheet.append(line[0])                # 加入非疑问词列表
                    # print(line[0], '... 不符合关键词要求')
                index += 1

        csv_reader = None
        del csv_reader                                           # 删除变量

        endTime = time()                                         # 运行结束时间

        print('筛选完毕...用时', '%.6fs' % (endTime - startTime))
        print('关键词总量:', len(rawSheet), '\t疑问词:', len(filterSheet), '\t非疑问词:', len(nofilterSheet))

        print('\n正在写入...')

        start2Time = time()

        sheetNames = ['未处理', '疑问词', '非疑问词']

        # 创建workbook
        workbook = Workbook()
        # 创建worksheet
        worksheet1 = workbook.create_sheet(index=0, title=sheetNames[0])
        worksheet2 = workbook.create_sheet(index=1, title=sheetNames[1])
        worksheet3 = workbook.create_sheet(index=2, title=sheetNames[2])

        i = 1
        for line in rawSheet:                        # 未处理sheet
            # 写入excel
            # 参数对应 行, 列, 值
            worksheet1.cell(row=i, column=1).value = str(line)
            i += 1
        rawSheet = None
        del rawSheet 

        print('已写入...未处理Sheet')

        i = 1
        for line in filterSheet:                     # 疑问词sheet
            worksheet2.cell(row=i, column=1).value = line
            i += 1
        filterSheet = None
        del filterSheet

        print('已写入...疑问词Sheet')

        i = 1
        for line in nofilterSheet:                   # 非疑问词sheet
            worksheet3.cell(row=i, column=1).value = line
            i += 1
        nofilterSheet = None
        del nofilterSheet

        print('已写入...非疑问词Sheet')

        output = path.join(file['filedir'], file['filename']) + '[已处理].xlsx'

        print('正在保存文件...')

        # 保存
        workbook.save(output)

        endTime = time()

        print('保存完毕...用时', '%.6fs\n' % (endTime - start2Time))

        worksheet1 = worksheet2 = worksheet3 = workbook = None
        del worksheet1
        del worksheet2
        del worksheet3
        del workbook
        
        print('总耗时:', '%.6fs' % (endTime - startTime))
        print('保存路径:', output)
        collect()                                    # gc

#####################################################
# 开始执行
#####################################################
keys = readKeys()

while True:

    try:
        main()
    except KeyboardInterrupt as e:
        print('中止。')
    except Exception as e:
        print('出错:', e)